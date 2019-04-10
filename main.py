# -*- coding: utf-8 -*-
# author_=su
# time=2018.11.18

import re
import random
import json
import requests
import time
# import pymongo
from openpyxl import Workbook
from datetime import datetime
from bs4 import BeautifulSoup

#配置类
class Config:
    def __init__(self,proxyHost,proxyPort,proxyUser,proxyPass,cookie):
        self.proxyHost=proxyHost
        self.proxyPort=proxyPort
        self.proxyUser=proxyUser
        self.proxyPass=proxyPass
        self.cookie=cookie

    #代理设置
    def Proxies(self):
        proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
            "host": self.proxyHost,
            "port": self.proxyPort,
            "user": self.proxyUser,
            "pass": self.proxyPass,
        }
        proxies = {
            "http": proxyMeta,
            "https": proxyMeta,
        }
        return proxies

    # cookie设置
    def Headers(self):
        headers = {
            'accept': '*/*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            # 'cache-control': 'max-age=0',
            'cookie':self.cookie,
            # 'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
        }
        return headers

#获取信息类
class GetFormation:
    def __init__(self,oraginurl,count,headers,proxies):
        self.oraginurl=oraginurl
        self.headers=headers
        self.proxies=proxies
        self.count=count

    #主进程
    def Get_formation(self):
        print ('获取商品Id', self.count)
        memberid,htm=self.Getmemberid()
        process=ProcessFormation()
        if memberid!=0:
            print ('获取商品信息',self.count,memberid)
            Shop_dict,companyname,productname=self.GetShop(htm)
            print ('获取商品评价信息',self.count)
            Review_Temporary=self.GetReviewers(companyname,productname,memberid)
            print ('获取商品物流信息',self.count)
            Trtanction_Temporary=self.GetTranction(companyname,productname,memberid)
            process.Save('success.txt',self.oraginurl)
        else:
            print('失败',self.count)
            process.Save('failed.txt',self.oraginurl)  #收集失败url
            Shop_dict={}
            Review_Temporary=[]
            Trtanction_Temporary=[]
        return Shop_dict,Review_Temporary,Trtanction_Temporary

    #获取店铺id  提供给api构造url
    def Getmemberid(self):
        try:
            htm = requests.get(self.oraginurl, headers=self.headers, proxies=self.proxies).text
            time.sleep(random.randint(1, 2))
        except:
            try:
                htm = requests.get(self.oraginurl, headers=self.headers, proxies=self.proxies).text
            except:
                htm = ''
                memberid = 0
                return memberid,htm
        # memberid = htm[htm.find('memberId":"') + len('memberId":"'): htm.find('",', htm.find('memberId":"'))]
        try:
            memberid=re.findall(r"member_id:\"(.*?)\"",htm)[0]
        except:
            memberid = 0
        return memberid, htm

    #采集店铺描述信息
    def GetShop(self,html):
        soup = BeautifulSoup(html,"html.parser")
        try: #产品名称
            productname = soup.find(id='mod-detail-title').find('h1').text
        except:
            productname = ''
        try: #公司名称
            companyname = soup.find(class_='company-name').text
        except:
            companyname = ''
        try:#公司年限
            companyyear=soup.find(class_='tp-year').text
        except:
            companyyear=''
        try: #价格 !
            price=''
            price_list=soup.find_all(class_='price-text')
            for m in price_list:
                price=price + m.text
        except:
            price=''
        try: #30天成交量
            bargain_number=soup.find(class_='bargain-number').find(class_="value").text
        except:
            bargain_number=''
        try: #30天内评价量
            satisfaction_number = soup.find(class_="satisfaction-number").find(class_='value').text
        except:
            satisfaction_number=''
        try: #支付方式
            pay_style=soup.find(class_='tradeway-container').text.replace('\n','')
        except:
            pay_style=''
        try: #货描
            description=soup.find(class_='topbar-bsr').find(class_='description-value-higher-hm').text
        except:
            description=''
        try:#响应
            response=soup.find(class_='topbar-bsr').find(class_='description-value-higher-xy').text
        except:
            response = ''
        try: #发货
            speed=soup.find(class_='topbar-bsr').find(class_='description-value-higher-fh').text
        except:
            speed=''
        try:#回头率
            back=soup.find(class_='topbar-bsr').find(class_='description-value-ht').text
        except:
            back=''
        try: #累积加工件数 交期准确率 货品满意度  !
            trade_list =soup.find(class_='mod-detail-processCustom-trade-preview').attrs['data-mod-config']
            trade_list = json.loads(trade_list)
            accumulativeProcessAmount=trade_list['accumulativeProcessAmount']
            deliveryPrecisionRatio=trade_list['deliveryPrecisionRatio']
            averageStarLevel=trade_list['averageStarLevel']
            if deliveryPrecisionRatio=='':
                deliveryPrecisionRatio='--'
            if accumulativeProcessAmount=='0':
                accumulativeProcessAmount='--'
        except:
            accumulativeProcessAmount=''
            deliveryPrecisionRatio = ''
            averageStarLevel = ''

        #获取动态加载数据
        pattern = re.compile(r'\d+')
        itemid = re.findall(pattern, self.oraginurl)[1]
        pageid=re.findall(r"pageid':'laputa(\d+)'",html)
        pageid='laputa'+str(pageid)
        active_url='https://laputa.1688.com/offer/ajax/widgetList.do?sk=processing&callback=jQuery17205061757331833663_1547898215182&blocks=&data=offerdetail_version2018_report%2Cofferdetail_version2018_serviceDesc%2Cofferdetail_ditto_postage%2Cofferdetail_ditto_offerSatisfaction%2Cofferdetail_w1190_guarantee%2Cofferdetail_w1190_tradeWay%2Cofferdetail_ditto_whosaleself&offerId='+itemid+'&pageId='+pageid
        html2 = requests.get(active_url, headers=self.headers).text
        #time.sleep(random.randint(2, 5))
        try:#可跟踪物流比
            delivery_percent = re.findall(r"logisticsDetailRatio\":(.*?),", html2)[0]
        except:
            delivery_percent=''
        try:
            delivery_time=re.findall(r"averageDeliverTimes\":(.*?),", html2)[0]
        except: #发货时长
            delivery_time=''
        try: # 实力保障
            stength = ''
            stength_list=re.findall(r"serviceName\":(.*?),", html2)
            for m in stength_list:
                stength = stength+m+','
            stength=stength.replace('}','').replace('"','').replace(']','')+'买家保障'
        except:
            stength=''
        try: #快递费
            cost=re.findall(r"costItems(.*?)value\":(\d+)", html2)[0][1]
        except:
            cost='卖家承担运费'
        try: #交易支持
            support_style=re.findall(r"offerdetail_w1190_tradeWay(.*)names\":(.*?)],", html2)[0][1] + ',"混批"]'
        except:
            support_style=''
        try: #出货 !  reserveRange
            manufacture_number_list=re.findall(r"reserveRange(.+?)]",html2)[0].replace('\\','').strip('"').strip(':')+']'
            manufacture_number_list=eval(manufacture_number_list)
            manufacture=''
            for m in manufacture_number_list:
                amount='数量(件): >'+str(m['beginAmount'])
                day='预计时间(天): '+str(m['date'])
                manufacture=manufacture+amount+day+'\n'
        except:
            manufacture=''
        information={}
        information['companyname']=companyname
        information['companyyear'] = companyyear
        information['productname'] = productname
        information['price'] = price
        information['cost'] = cost
        information['delivery_time'] = delivery_time
        information['delivery_percent'] = delivery_percent
        information['bargain_number'] = bargain_number
        information['satisfaction_number'] = satisfaction_number
        information['stength'] = stength
        information['pay_style'] = pay_style
        information['support_style'] = support_style
        information['manufacture_number']=manufacture
        information['description'] = description
        information['response'] = response
        information['speed'] = speed
        information['back'] = back
        information['countnumber'] = accumulativeProcessAmount
        information['accurate'] =deliveryPrecisionRatio
        information['product_satisfaction'] = averageStarLevel
        information['link']=self.oraginurl
        return information,companyname,productname

    #采集评价信息
    def GetReviewers(self,companyname,productname,memberid):
        pattern = re.compile(r'\d+')
        itemid = re.findall(pattern, self.oraginurl)[1]
        Review_Temporary=[]
        for i in range(1, 100):  # 每页15条
            url = 'https://rate.1688.com/remark/offerDetail/rates.json?_input_charset=GBK&offerId=' + str(
                itemid) + '&page=' + str(
                i) + '&pageSize=15&starLevel=&orderBy=&semanticId=&showStat=0&content=1&t=1542595638441&memberId=' + str(memberid)
            try:
                html = requests.get(url, headers=self.headers, proxies=self.proxies).text
            except:
                try:
                    html = requests.get(url, headers=self.headers, proxies=self.proxies).text
                except:
                    break
            try:
                data = json.loads(html)
                lists = data['data']['rates']
            except:
                break
            if len(lists)==0:
                break
            else:
                for li in lists:
                    option = {}
                    try:
                        option['number'] = self.count
                        option['company_name'] = companyname  # 公司名称
                        option['product_name'] = productname  # 产品名称
                        option['link'] = oraginurl  # 网址
                        option['star_level'] = li['rateItem'][0]['starLevel']  # 星级
                        option['review_content'] = li['rateItem'][0]['remarkContent']  # 评价内容
                        option['amount'] = li['quantity']  # 数量
                        option['all_amount'] = li['countQuantity']  # 累积采购
                        Review.append(option)
                        Review_Temporary.append(option)
                    except Exception as e:
                        print(e)
                        break
        return Review_Temporary

    #采集物流信息
    def GetTranction(self,companyname,productname,memberid):
        pattern = re.compile(r'\d+')
        itemid = re.findall(pattern, self.oraginurl)[1]
        Trtanction_Temporary=[]
        for i in range(1, 100):  # 每次可以采集20条
            url = 'https://rate.1688.com/remark/offerDetail/saleRecords.json?offerId=' + str(
                itemid) + '&currentPage='+str(i)+'&showStat=1&rateContentType=&memberId=' + str(memberid) + '&recordStatus=0'
            try:
                html = requests.get(url, headers=self.headers, proxies=self.proxies).text
            except:
                try:
                    html = requests.get(url, headers=self.headers, proxies=self.proxies).text
                except:
                    break
            try:
                data = json.loads(html)
                lists = data['data']['orderDisplayEntryList']
            except:
                break
            if len(lists)==0:
                break
            for n in lists:
                tranction_dict = {}
                tranction_dict['number']=self.count
                tranction_dict['company_name'] = companyname  # 公司名称
                tranction_dict['product_name'] = productname  # 产品名称
                tranction_dict['link'] = oraginurl
                tranction_dict['purchase_repeat'] = data['data']['offerSaleRecordStat']['repeatBuyCount']  # 重复购买率
                tranction_dict['purchase_man_number'] = data['data']['offerSaleRecordStat']['buyerTotal']  # 采购人数
                tranction_dict['pagecount'] = data['data']['totalCount']  # 物流总条数
                number = n['quantity']  # 用户采购数量
                try:
                    totalnumber = n['countBuyerQuantity']  # 用户历史采购总量
                except:
                    totalnumber = '0'
                tranction_dict['purchase_number'] = str(number) + '件累计采购' + str(totalnumber)  # 采购数量结合
                tranction_dict['purchase_businessman'] = n['buyerName']  # 采购商昵称
                tranction_dict['purchase_time'] = n['buyerPayTime']  # 采购时间
                try:
                    tranction_dict['purchase_style'] = str(n['specInfo'])  # 商品规格
                except:
                    tranction_dict['purchase_style'] = ''
                Trtanction_Temporary.append(tranction_dict)
        return Trtanction_Temporary

#处理信息类
class ProcessFormation:

    def GetUrl(self,path):
        url_list=[]
        with open(path,'r') as f:
            temporary_list=f.readlines()
        for url in temporary_list:
            if url not in url_list:
                url=url.replace(',','').replace('\n','').replace('[','').replace(']','')
                url_list.append(url)
        return url_list

    def Save(self,path,data):
        with open(path,'a+',encoding='utf-8') as f:
            f.write(str(data)+'\n')

    # 写入excel
    def write_excel(self,list,event):
        wb = Workbook()
        ws = wb.active
        row = 2
        col = 1
        time = str(datetime.now().date())
        if event == 'aggregate':
            ws['A1'] = '公司名称'
            ws['B1'] = '产品名称'
            ws['C1'] = '价格'
            ws['D1'] = '快递费'
            ws['E1'] = '实际发货时长'
            ws['F1'] = '可跟踪物流占比'
            ws['G1'] = '30天成交记录数量'
            ws['H1'] = '30天评论数量'
            ws['I1'] = '实力保障'
            ws['J1'] = '支付方式'
            ws['K1'] = '交易支持'
            ws['L1'] = '出货'
            ws['M1'] = '累积加工件数'
            ws['N1'] = '交期准确率'
            ws['O1'] = '货品满意度'
            ws['P1'] = '运营时间'
            ws['Q1'] = '货描'
            ws['R1'] = '响应'
            ws['S1'] = '发货'
            ws['T1'] = '回头率'
            ws['U1'] = '网址'
            for m in list:
                if len(m) !=0:
                    values = [m['companyname'], m['productname'], m['price'], m['cost'],
                              m['delivery_time'], m['delivery_percent'], m['bargain_number'], m['satisfaction_number'],
                              m['stength'], m['pay_style'], m['support_style'], m['manufacture_number'],m['countnumber'],m['accurate'],
                              m['product_satisfaction'], m['companyyear'], m['description'], m['response'], m['speed'],
                              m['back'], m['link']]
                    for i in values:
                        ws.cell(row=row, column=col, value=i)
                        col = col + 1
                    row = row + 1
                    col = 1
            filename = time + '-' + "Aggregate.xlsx"
            wb.save(filename)
        elif event == 'reviews':
            ws['A1'] = '序号'
            ws['B1'] = '网址'
            ws['C1'] = '公司名称'
            ws['D1'] = '产品名称'
            ws['E1'] = '星级'
            ws['F1'] = '数量'
            ws['G1'] = '累计采购'
            ws['H1'] = '评价'
            for m in list:
                values = [m['number'], m['link'], m['company_name'], m['product_name'], m['star_level'], m['amount'],
                          m['all_amount'], m['review_content']]
                for i in values:
                    i=str(i).replace('\x14','')
                    ws.cell(row=row, column=col, value=i)
                    col = col + 1
                col = 1
                row = row + 1
            filename = time + '-' + "review.xlsx"
            wb.save(filename)
        else:
            ws['A1'] = '序号'
            ws['B1'] = '网址'
            ws['C1'] = '公司名称'
            ws['D1'] = '产品名称'
            ws['E1'] = '采购重复率'
            ws['F1'] = '采购人数'
            ws['G1'] = '总共条数'
            ws['H1'] = '采购商'
            ws['I1'] = '商品规格'
            ws['J1'] = '数量'
            ws['K1'] = '成交时间'
            for m in list:
                values = [m['number'], m['link'], m['company_name'], m['product_name'], m['purchase_repeat'],
                          m['purchase_man_number'], m['pagecount'], m['purchase_businessman'], m['purchase_style'],
                          m['purchase_number'], m['purchase_time']]
                for i in values:
                    ws.cell(row=row, column=col, value=i)
                    col = col + 1
                col = 1
                row = row + 1
            filename = time + '-' + "transaction.xlsx"
            wb.save(filename)

if __name__ == '__main__':
    #代理设置
    proxyHost = "http-dyn.abuyun.com"
    proxyPort = "9020"
    proxyUser = "H3EY606Y4F5D2H0D"      #修改     1
    proxyPass = "F20D99D586BF789E"      #修改     2
    #cookie设置                            #修改     3
    cookies='UM_distinctid=1683212fbde14-0ace1ac1872f65-b781636-1fa400-1683212fbdf3c; ali_ab=61.170.164.58.1547028334711.5; cna=brW8FCOD7EsCAT2qpDqg8EdZ; l=aBtbS_GpyYwSnHLmQMai56m76707StfPt2KX1MaHVTEhNOMG7RXy1Jro-VwWT_qC5f4O_K-5F; cookie2=1ccb279d00c2ccffee5b5bc7201f1e75; t=98c32eb229681af590101fa1d965e4b9; _tb_token_=3e154f330dfee; ali_apache_tracktmp=c_w_signed=Y; __rn_alert__=false; cookie1=UIMaxdIwIIRPeh4fjQw9wpWD7tpV8ZbcNk2%2BRWdyh5c%3D; cookie17=UU8IOfoDXDu4QQ%3D%3D; sg=t33; csg=0884280f; lid=%E5%AE%9D%E5%AE%9D%E6%9C%80%E7%88%B1%E7%9A%84%E4%BA%BAqt; unb=2705262233; __cn_logon__=true; __cn_logon_id__=%E5%AE%9D%E5%AE%9D%E6%9C%80%E7%88%B1%E7%9A%84%E4%BA%BAqt; ali_apache_track=c_mid=b2b-27052622334a5a0|c_lid=%E5%AE%9D%E5%AE%9D%E6%9C%80%E7%88%B1%E7%9A%84%E4%BA%BAqt|c_ms=1; _nk_=%5Cu5B9D%5Cu5B9D%5Cu6700%5Cu7231%5Cu7684%5Cu4EBAqt; last_mid=b2b-27052622334a5a0; _csrf_token=1548319073143; _is_show_loginId_change_block_=b2b-27052622334a5a0_false; _show_force_unbind_div_=b2b-27052622334a5a0_false; _show_sys_unbind_div_=b2b-27052622334a5a0_false; _show_user_unbind_div_=b2b-27052622334a5a0_false; alicnweb=homeIdttS%3D52558415885919043677110903947882455091%7Ctouch_tb_at%3D1548318884988%7Clastlogonid%3D%25E5%25AE%259D%25E5%25AE%259D%25E6%259C%2580%25E7%2588%25B1%25E7%259A%2584%25E4%25BA%25BAqt%7ChomeIdttSAction%3Dtrue%7Chp_newbuyerguide%3Dtrue; x5sec=7b22726174653b32223a223330346431616339383034623065636139373738626337626231386466383361434d583270654946454a3754392b72657a37376b74414561444449334d4455794e6a49794d7a4d374d513d3d227d; isg=BBkZLyCGocEcJn0pbA81WVRGKAUzDgcr_3_UezvPbcB-Qj3Ug_NXKEUUQEaReqWQ'
    count=1
    #实例化配置项
    config=Config(proxyHost,proxyPort,proxyUser,proxyPass,cookies)
    process=ProcessFormation()
    proxies=config.Proxies()
    headers=config.Headers()
    url_list=process.GetUrl('url.txt')           #修改要爬取的url集合的路径     4

    success = process.GetUrl('success.txt')
    new_success = []
    for url in success:
        url = url.strip('\n')
        new_success.append(url)

    Review=[]
    Trtanction=[]
    Aggregate=[]
    for oraginurl in url_list:
        #去除已经爬取的url
        oraginurl=oraginurl.replace('\'','').replace(',','').strip('\n')
        if oraginurl in new_success:
            print("Url" + str(count) + " already exists.")
            count = count + 1
            continue
        getformation=GetFormation(oraginurl,count,headers,proxies)
        Shop_dict,  Review_Temporary, Trtanction_Temporary=getformation.Get_formation()
        #存入临时文件txt #存入列表
        if Shop_dict!='':
            process.Save('aggregate.txt', Shop_dict)
            Aggregate.append(Shop_dict)
        if len(Review_Temporary)!=0:
            process.Save('reviews.txt', Review_Temporary)
            Review.append(Review_Temporary)
        if len(Trtanction_Temporary) != 0:
            process.Save('transaction.txt', Trtanction_Temporary)
            Trtanction.append(Trtanction_Temporary)
        count=count+1
    process.write_excel(Aggregate,'aggregate')
    process.write_excel(Review, 'reviews')
    process.write_excel(Trtanction, 'transaction')

